import numpy as np
from numpy import square, arctan2, cos, sin, pi, sqrt, degrees
from numpy.linalg import norm, inv

import pandas as pd

from scipy import optimize
from scipy.spatial import KDTree
from scipy.spatial.distance import cdist

from openpyxl import load_workbook
from collections import defaultdict

from copy import copy

import numbers, fcl

# local libraries
from utils.well_mesh import WellMesh
from utils.well_scene import *
from utils.well_collision import *
from utils.well_utils import *

def get_standard_data(filename):
    # import data from Excel
    workbook = load_workbook(filename, data_only=True)
    sheets = workbook.sheetnames

    # extract the sheets with the input data
    wells = []
    wells.extend([well for well in list(sheets) if well.split()[-1] == "well"])

    data = {
        "acr": {},
        "wells": {},
    }

    for i, well in enumerate(wells):
        sheet = workbook[well]
        
        if well == "Reference well":
            data = acr_setup(sheet, data)
        
        data = get_data(well, sheet, data)

    return data

def get_data(well, sheet, data):
    temp = dict(
        MD = [],
        IncDeg = [],
        AziDeg = [],
        TVD = [],
        N = [],
        E = [],
        sigmaH = [],
        sigmaL = [],
        sigmaA = [],
    )

    for row in sheet.iter_rows(
        min_row=17,
        max_row=(),
        min_col=2,
        max_col=10,
    ):
        if isinstance(row[0].value, numbers.Real):
            temp["MD"].append(row[0].value)
            temp["IncDeg"].append(row[1].value)
            temp["AziDeg"].append(row[2].value)
            temp["TVD"].append(row[3].value)
            temp["N"].append(row[4].value)
            temp["E"].append(row[5].value)
            temp["sigmaH"].append(row[6].value)
            temp["sigmaL"].append(row[7].value)
            temp["sigmaA"].append(row[8].value)

    data["wells"][f"{well}"] = temp

    return data

def acr_setup(sheet, data):
    data["acr"]["Sm"] = sheet["I4"].value
    data["acr"]["sigmapa"] = sheet["I5"].value
    data["acr"]["k"] = sheet["I6"].value
    data["acr"]["reference_h_and_c"] = sheet["I7"].value
    data["acr"]["offset_h_and_c"] = sheet["I8"].value

    return data

def make_meshes(data):
    k = data["acr"]["k"]
    Sm = data["acr"]["Sm"]
    # Sm = 1
    k = 3

    for well in data["wells"]:
        survey = np.array([data["wells"][well]["MD"], data["wells"][well]["IncDeg"], data["wells"][well]["AziDeg"]]).T
        NEVs = np.array([data["wells"][well]["N"], data["wells"][well]["E"], data["wells"][well]["TVD"]]).T

        if well == "Reference well":
            radius = data["acr"]["reference_h_and_c"]
            Sm = 0
        else:
            radius = data["acr"]["offset_h_and_c"]
        radius /= (2 * 100) / 2.54
        
        h = np.array(data["wells"][well]["sigmaH"])
        l = np.array(data["wells"][well]["sigmaL"])
        a = np.array(data["wells"][well]["sigmaA"])

        cov = np.array([
            [square(h), h * l, h * a],
            [h * l, square(l), l * a],
            [h * a, l * a, square(a)]
        ]).T

        wm = WellMesh(
            survey,
            NEVs,
            cov,
            cov_HLAs=True,
            N_verts=12,
            sigma=k,
            wellbore_radius=radius,
            surface_margin=Sm,
            degrees=True,
            poss=False,
        )

        data["wells"][well]["mesh"] = wm

    return data

def make_survey(data, well):
    start_nev = data["wells"][offset]["mesh"].NEVs[0]
    y, x, z = start_nev
    start_xyz = np.array([x, y, z])
    return Survey(
        md=data["wells"][well]["MD"],
        inc=data["wells"][well]["IncDeg"],
        azi=data["wells"][well]["AziDeg"],
        start_nev=start_nev,
        start_xyz=start_xyz
    )


def interpolate_sigmas(data, i, mult):
    sigmaH_new = (
        data["wells"][offset]["sigmaH"][i - 1]
        + mult * (data["wells"][offset]["sigmaH"][i] - data["wells"][offset]["sigmaH"][i - 1])
    )
    sigmaL_new = (
        data["wells"][offset]["sigmaL"][i - 1]
        + mult * (data["wells"][offset]["sigmaL"][i] - data["wells"][offset]["sigmaL"][i - 1])
    )
    sigmaA_new = (
        data["wells"][offset]["sigmaA"][i - 1]
        + mult * (data["wells"][offset]["sigmaA"][i] - data["wells"][offset]["sigmaA"][i - 1])
    )
    return (sigmaH_new, sigmaL_new, sigmaA_new)


# def fun(x, survey, index, station):
#     temp = interpolate_survey(survey, index, x[0])
#     dist = norm(temp["poss"] - station)
#     return dist

def fun(x, survey, index, station):
    s = interpolate(survey, index, x[0])
    new_pos = np.array([s.n, s.e, s.tvd]).T[1]
    dist = norm(new_pos - station, axis=-1)

    return dist


def get_standard_separation(data, reference, offset, kop_index=0):
    idx = np.argmin(
        cdist(
            data["wells"][reference]["mesh"].NEVs, data["wells"][offset]["mesh"].NEVs
        ), axis=-1
    )
    s_offset = make_survey(data, offset)

    reference_poss = data["wells"][reference]["mesh"].NEVs
    closest = []
    for j, (i, station) in enumerate(zip(idx, reference_poss)):
        if j < kop_index: continue
        if i > 0:
            bnds = [(0, data["wells"][offset]["MD"][i] - data["wells"][offset]["MD"][i - 1])]
            res_1 = optimize.minimize(
                fun,
                bnds[0][1],
                method='SLSQP',
                bounds=bnds,
                args=(s_offset, i-1, station)
                )
            mult = res_1.x[0] / (bnds[0][1] - bnds[0][0])
            sigma_new_1 = interpolate_sigmas(data, i, mult)
        else: res_1 = False

        if i < len(data["wells"][offset]["mesh"].NEVs) - 1:
            bnds = [(0, data["wells"][offset]["MD"][i + 1] - data["wells"][offset]["MD"][i])]
            res_2 = optimize.minimize(
                fun,
                bnds[0][0],
                method='SLSQP',
                bounds=bnds,
                args=(s_offset, i, station)
                )
            mult = res_2.x[0] / (bnds[0][1] - bnds[0][0])
            sigma_new_2 = interpolate_sigmas(data, i + 1, mult)
        else: res_2 = False

        if res_1 and res_2 and res_1.fun < res_2.fun or not res_2:
            closest.append((station, interpolate(s_offset, i - 1, res_1.x[0]), res_1, sigma_new_1))
        else:
            closest.append((station, interpolate(s_offset, i, res_2.x[0]), res_2, sigma_new_2))



    # offset_MDs = np.array([r[1]["mds"] for r in closest])
    offset_MDs = np.array([r[1].md[1] for r in closest])
    offset_survey = np.array([r[1].survey_rad[1] for r in closest])
    # offset_survey = np.array([
    #     np.hstack(data["wells"][well]["MD"]),
    #     radians(np.hstack(data["wells"][well]["IncDeg"])),
    #     radians(np.hstack(data["wells"][well]["AziDeg"]))
    # ]).T[idx]
    n, e, tvd = np.array([[r[1].n[1], r[1].e[1], r[1].tvd[1]] for r in closest]).T
    offset_sigmaH = np.array([r[3][0] for r in closest])
    offset_sigmaL = np.array([r[3][1] for r in closest])
    offset_sigmaA = np.array([r[3][2] for r in closest])
    # offset_sigmaH = np.hstack(data["wells"][well]["sigmaH"])[idx]
    # offset_sigmaL = np.hstack(data["wells"][well]["sigmaL"])[idx]
    # offset_sigmaA = np.hstack(data["wells"][well]["sigmaA"])[idx]
    offset_NEVs = np.array([n, e, tvd]).T
    # offset_NEVs = np.array([
    #     np.array(data["wells"][well]["N"])[idx],
    #     np.array(data["wells"][well]["E"])[idx],
    #     np.array(data["wells"][well]["TVD"])[idx]
    # ]).T[kop_index:]
    # offset_vecs = np.vstack([r[1]["vecs"] for r in closest])
    
    # need to convert NEV to HLA
    ref_delta_NEVs = offset_NEVs - data["wells"][reference]["mesh"].NEVs[kop_index:]
    unit_ref_delta_NEVs = ref_delta_NEVs / norm(ref_delta_NEVs, axis=-1).reshape(-1,1)

    off_delta_NEVs = data["wells"][reference]["mesh"].NEVs[kop_index:] - offset_NEVs
    unit_off_delta_NEVs = off_delta_NEVs / norm(off_delta_NEVs, axis=-1).reshape(-1,1)
    offset_delta_HLAs = np.stack([
        NEV_to_HLA(s.reshape(-1,3), nev.reshape(-1,3), cov=False)
        for s, nev in zip(offset_survey, off_delta_NEVs)
    ]).reshape(-1,3)

    # get the unit [H, L, A] vectors
    unit_off_HL_vec = offset_delta_HLAs / norm(offset_delta_HLAs, axis=-1).reshape(-1,1)

    # try pedal curve
    off_PCR = []
    off_H = offset_sigmaH
    off_L = offset_sigmaL
    for vec, cov in zip(
        unit_off_HL_vec[:,:2] / norm(unit_off_HL_vec[:,:2], axis=-1).reshape(-1,1),
        # off_delta_NEVs[:,:2] / norm(off_delta_NEVs[:,:2], axis=-1).reshape(-1,1),
        np.array([
            [off_H ** 2, off_H * off_L],
            [off_H * off_L, off_L ** 2]
        ]).T
    ):
        off_PCR.append(sqrt(np.dot(np.dot(vec, cov), vec.T)))

    # need to convert NEV to HLA
    reference_delta_HLAs = np.vstack([
        NEV_to_HLA(s.reshape(-1,3), nev.reshape(-1,3), cov=False)
        for s, nev in zip(
            data["wells"][reference]["mesh"].survey[kop_index:], unit_ref_delta_NEVs)
    ]).reshape(-1,3)
    # ref_hoz_bearing = arctan2(reference_delta_HLAs[:,1], reference_delta_HLAs[:,0])

    # get the unit [H, L] vectors (could just inverse the offset one?)
    unit_ref_HL_vec = reference_delta_HLAs / norm(reference_delta_HLAs, axis=-1).reshape(-1,1)

    ref_PCR = []
    ref_H = np.array(data["wells"][reference]["sigmaH"])[kop_index:]
    ref_L = np.array(data["wells"][reference]["sigmaL"])[kop_index:]
    ref_A = np.array(data["wells"][reference]["sigmaA"])[kop_index:]
    for vec, cov in zip(
        unit_ref_HL_vec, # [:,:2] /  norm(unit_ref_HL_vec[:,:2], axis=-1).reshape(-1,1),
        # ref_delta_NEVs[:,:2] / norm(ref_delta_NEVs[:,:2], axis=-1).reshape(-1,1),
        np.array([
            [ref_H ** 2, ref_H * ref_L, ref_H * ref_L],
            [ref_H * ref_L, ref_L ** 2, ref_L * ref_A],
            [ref_H * ref_A, ref_L * ref_A, ref_A ** 2]
        ]).T
    ):
        ref_PCR.append(sqrt(np.dot(np.dot(vec, cov), vec.T)))

    k = data["acr"]["k"]
    sigmapa = data["acr"]["sigmapa"]
    Sm = data["acr"]["Sm"]
    Rr = data["acr"]["reference_h_and_c"]
    Ro = data["acr"]["offset_h_and_c"]
    calc_hole = (Rr + Ro) * 2.54 / 200

    dist_CC_Clr = norm(ref_delta_NEVs, axis=1)

    hoz_bearing = arctan2(ref_delta_NEVs[:,1], ref_delta_NEVs[:,0])

    zeros = np.zeros_like(offset_sigmaH)
    # off_cov = np.array([
    #     [offset_sigmaH ** 2, offset_sigmaH * offset_sigmaL, offset_sigmaH * offset_sigmaA],
    #     [offset_sigmaH * offset_sigmaL, offset_sigmaL ** 2, offset_sigmaL * offset_sigmaA],
    #     [offset_sigmaH * offset_sigmaA, offset_sigmaL * offset_sigmaA, offset_sigmaA ** 2]
    # ]).T
    # off_cov = np.array([
    #     [offset_sigmaH ** 2, offset_sigmaH * offset_sigmaL, zeros],
    #     [offset_sigmaH * offset_sigmaL, offset_sigmaL ** 2, zeros],
    #     [zeros, zeros, zeros]
    # ]).T
    off_cov = np.array([
        [offset_sigmaH ** 2, zeros, zeros],
        [zeros, offset_sigmaL ** 2, zeros],
        [zeros, zeros, offset_sigmaA ** 2]
    ]).T

    off_vecs = []
    # off_NEV_vec = off_delta_NEVs / norm(off_delta_NEVs, axis=-1).reshape(-1,1)
    for vec, bearing in zip(unit_off_HL_vec, hoz_bearing):
        # if abs(bearing) <= pi/2:
        if vec[2] > -1: #-0.25:
            temp = [vec[0], vec[1], 0]
            # temp /= norm(temp)
        else:
            # temp = vec
            if abs(vec[0]) > abs(vec[1]):
                temp = [vec[0], 0, vec[2]]
            else:
                temp = [0, vec[1], vec[2]]
            temp /= norm(temp)
        off_vecs.append(temp)
    off_vecs = np.vstack(off_vecs)

    off_PCR = [sqrt(np.dot(np.dot(vec, cov), vec.T)) for vec, cov in zip(unit_off_HL_vec, off_cov)]
    # off_PCR = np.hstack([sqrt(np.dot(np.dot(vec, cov), vec.T)) for vec, cov in zip(off_vecs, off_cov)])

    zeros = np.zeros_like(ref_H)
    ref_cov = np.array([
        [ref_H ** 2, zeros, zeros],
        [zeros, ref_L ** 2, zeros],
        [zeros, zeros, ref_A ** 2]
    ]).T
    # ref_cov = np.array([
    #     [ref_H ** 2, ref_H * ref_L, ref_H * ref_A],
    #     [ref_H * ref_L, ref_L ** 2, ref_L * ref_A],
    #     [ref_H * ref_A, ref_L * ref_A, ref_A ** 2]
    # ]).T

    ref_vecs = []
    for vec, bearing in zip(unit_ref_HL_vec, hoz_bearing):
        # if abs(bearing) <= pi/2:
        if vec[2] > -0.25:
            temp = [vec[0], vec[1], 0]
            temp /= norm(temp)
        else:
            # temp = vec
            # temp = [0, vec[1], vec[2]]
            if abs(vec[0]) > abs(vec[1]):
                temp = [vec[0], 0, vec[2]]
            else:
                temp = [0, vec[1], vec[2]]
            temp /= norm(temp)
        ref_vecs.append(temp)
    ref_vecs = np.vstack(off_vecs)

    ref_PCR = [sqrt(np.dot(np.dot(vec, cov), vec.T)) for vec, cov in zip(unit_ref_HL_vec, ref_cov)]
    # ref_PCR = [sqrt(np.dot(np.dot(vec, cov), vec.T)) for vec, cov in zip(ref_vecs, ref_cov)]

    theta = pi - arctan2(unit_ref_HL_vec[:,0], unit_ref_HL_vec[:,1])
    r = sqrt(
        (ref_L ** 2) * (cos(theta) ** 2) + (ref_H ** 2) * (sin(theta) ** 2)
    )
    # r = ref_H * cos(theta/2) ** 2
    a = ref_L[-1]
    b = ref_H[-1]
    x = np.linspace(0,1,10)
    y = np.linspace(0,1,10)
    x,y = np.meshgrid(x,y)
    theta = np.linspace(0, 2*pi, 100)
    x = ref_L[-1] ** 2 * cos(theta) ** 2
    y = ref_H[-1] ** 2 * sin(theta) ** 2
    

    # import plotly.graph_objects as go
    # fig = go.Figure(data=go.Scatter(
    #     x=x,
    #     y=y,
    # ))
    # fig.show(renderer="iframe")

    sigmaS = sqrt((np.hstack(ref_PCR) ** 2) + (np.hstack(off_PCR) ** 2))

    SF = (dist_CC_Clr - calc_hole - Sm) / (k * sqrt((sigmaS ** 2) + (sigmapa ** 2)))
    # SF = (dist_CC_Clr - calc_hole - Sm) / (k * (ref_PCR + off_PCR + sigmapa))

    

    data["wells"][offset]["results"].update(
        off_md = list(offset_MDs),
        off_tvd = list(tvd),
        off_n = list(n),
        off_e = list(e),
        # off_sigmaH = list(offset_sigmaH),
        # off_sigmaL = list(offset_sigmaL),
        # off_sigmaA = list(offset_sigmaA),
        hoz_bearing = list((degrees(hoz_bearing) + 360) % 360),
        dist_CC_Clr = list(dist_CC_Clr),
        ref_PCR = list(ref_PCR),
        off_PCR = list(off_PCR),
        calc_hole = calc_hole,
        ref_vec_H = list(unit_ref_HL_vec[:,0]),
        ref_vec_L = list(unit_ref_HL_vec[:,1]),
        ref_vec_A = list(unit_ref_HL_vec[:,2]),
        off_vec_H = list(unit_off_HL_vec[:,0]),
        off_vec_L = list(unit_off_HL_vec[:,1]),
        off_vec_A = list(unit_off_HL_vec[:,2]),
        # sigmaS = list(sigmaS),
        # d = (list(d)),
        ISCWSA_ACR = list(SF)
    )



    return data



    print("Hello world")


if __name__ == "__main__":

    data = get_standard_data(filename=f"reference/standard-set-of-wellpaths-for-evaluating-clearance-scenarios-r4-17-may-2017.xlsm")

    s = Survey(
        md = data["wells"]["Reference well"]["MD"],
        inc = data["wells"]["Reference well"]["IncDeg"],
        azi = data["wells"]["Reference well"]["AziDeg"],
    )

    # s_int = interpolate(s, 75, 10)

    # s_int = interpolate_survey(s, 75, step=20)

    data = make_meshes(data)

    data = make_trimesh_scene(data)

    # save the CollisionManager to file
    data["scene"].export("data/scene.glb")

    # make a version for visualization and save it
    scene_transform = transform_trimesh_scene(data["scene"], origin=([0,0,0]), scale=100, redux=1)
    scene_transform.export("blender/scene_transform.glb")

    data = make_collision_manager(data)

    # for each reference offset tuple we'll need to step through the survey
    for well in data["wells"].keys():
        if well == "Reference well": continue
        reference = "Reference well"
        offset = well

        cm_offset = trimesh.collision.CollisionManager()
        cm_offset.add_object(offset, data["wells"][offset]["mesh"].mesh)

        reference_survey = data["wells"][reference]["mesh"].survey
        reference_NEVs = data["wells"][reference]["mesh"].NEVs
        reference_covs = copy(data["wells"][reference]["mesh"].cov_HLAs.T)
        reference_radius = data["acr"]["reference_h_and_c"] / (2 * 100) * 2.54

        offset_NEVs = data["wells"][offset]["mesh"].NEVs

        SFs = [None]
        distances_CC = [None]
        separations = [None]
        collisions = [None]

        if offset == "10 - well":
            kod = data["wells"][well]["MD"][0]
            kod_reference_index = int(np.searchsorted(data["wells"][reference]["MD"], kod, side="left"))
            kod_reference_cov = reference_covs[kod_reference_index]
            # reference_cov_NEVs = HLA_to_NEV(reference_survey, reference_covs.T).T
            # kod_reference_cov_NEV = reference_cov_NEVs[kod_reference_index]
            # temp_cov_NEVs = np.zeros_like(reference_cov_NEVs[kod_reference_index:])
            # temp_cov_NEVs = (reference_cov_NEVs[kod_reference_index:].T - np.full_like(reference_cov_NEVs[kod_reference_index:], kod_reference_cov_NEV).T).T
            temp_covs = np.zeros_like(reference_covs[kod_reference_index:])
            temp_covs = (reference_covs[kod_reference_index:].T - np.full_like(reference_covs[kod_reference_index:], kod_reference_cov).T).T
            reference_covs = temp_covs
            reference_survey = reference_survey[kod_reference_index:]
            reference_NEVs = reference_NEVs[kod_reference_index:]

            mesh_relative = WellMesh(
                reference_survey,
                reference_NEVs,
                reference_covs,
                cov_HLAs=True,
                N_verts=12,
                sigma=3.0,
                wellbore_radius=reference_radius,
                surface_margin=0,
                degrees=False,
                poss=False,
            ).mesh
            
            scene_relative = trimesh.scene.scene.Scene()
            scene_relative.add_geometry(
                mesh_relative, node_name=reference, geom_name=reference, parent_node_name=None
                )
            scene_relative.add_geometry(
                data["wells"][offset]["mesh"].mesh, node_name=offset, geom_name=offset, parent_node_name=None
                )
            scene_relative_transform = transform_trimesh_scene(scene_relative, redux=1.0)
            scene_relative_transform.export("blender/scene_relative_transform.glb")
        

        for i, _  in enumerate(reference_survey):
            if i == 0: continue

            mesh_absolute = WellMesh(
                reference_survey[i-1:i+1],
                reference_NEVs[i-1:i+1],
                reference_covs[i-1:i+1],
                cov_HLAs=True,
                N_verts=12,
                sigma=3.0,
                wellbore_radius=reference_radius,
                surface_margin=0,
                degrees=False,
                poss=False,
            ).mesh

            ### using fcl direct ###
            creq = fcl.CollisionRequest(enable_contact=True)
            cres = fcl.CollisionResult()

            o1_name = reference
            o1 = fcl.CollisionObject(trimesh.collision.mesh_to_BVH(mesh_absolute))

            o2_name = well
            o2 = fcl.CollisionObject(trimesh.collision.mesh_to_BVH(data["wells"][offset]["mesh"].mesh))

            n_contacts = fcl.collide(o1, o2, creq, cres)

            dreq = fcl.DistanceRequest(enable_nearest_points=True)
            dres = fcl.DistanceResult()

            dist = fcl.distance(o1, o2, dreq, dres)

            # get collision results
            collision = cres.is_collision
            if collision:
                penetration = cres.contacts[0].penetration_depth
                separation = -penetration


            ###


            # to work out the SF, we need to figure out the closest points on the wellpath to the closest points
            # on the mesh
            distance_absolute = cm_offset.min_distance_single(mesh_absolute, return_name=True, return_data=True)
            closest_point_reference = distance_absolute[2].point("__external")
            name_offset_absolute = distance_absolute[1]
            closest_point_offset = distance_absolute[2].point(name_offset_absolute)
            reference_NEV = reference_NEVs[KDTree(reference_NEVs).query(closest_point_reference)[1]]

            offset_NEV = offset_NEVs[KDTree(
                offset_NEVs
            ).query(closest_point_offset)[1]]
            distance_CC = norm(offset_NEV - reference_NEV)
        
            # see if there's a collision   
            collision_absolute = cm_offset.in_collision_single(mesh_absolute, return_names=True, return_data=True)
            if collision_absolute[0]:
                # # needed to add "self._depth = contact.penetration_depth" to the ContactData class of trimesh.collision
                # # penetrations, points = np.vstack([[penetration._depth, penetration._point] for penetration in collision_absolute[2]]).T
                # penetrations, points = zip(*[[penetration._depth, penetration._point] for penetration in collision_absolute[2]])
                # penetrations = np.array(penetrations)
                # points = np.stack(points, axis=0)
                # try:
                #     closest_point_to_NEV = points[np.argmin(cdist(points, reference_NEV.reshape(-1,3))[np.where(penetrations > 0)])]
                # except: separation = 0
                # # closest_vertex_reference = np.asarray(mesh_absolute.vertices[np.argmin(cdist(mesh_absolute.vertices, closest_point_to_NEV.reshape(-1,3)))])
                # # separation = -norm(closest_vertex_reference - closest_point_to_NEV)
                # separation = -np.amax(penetrations)
                # # separation = -penetrations[np.argmin(cdist(points, reference_NEV.reshape(-1,3))[np.where(penetrations > 0)])][0]
                SF_absolute = distance_CC / (distance_CC - separation)

            else:
                separation = distance_absolute[0]
                SF_absolute = distance_CC / (distance_CC - separation)

            # update lists
            SFs.append(SF_absolute)
            distances_CC.append(distance_CC)
            separations.append(separation)
            # collisions.append(collision_absolute[0])
            collisions.append(collision)

        # get standard deviation
        # std_dev = get_standard_separation(data, reference, well)
        # mds = np.array([r[1]["mds"] for r in std_dev])
        # e, n, tvd = np.vstack([r[1]["poss"] for r in std_dev]).T

        # update well results
        data["wells"][well]["results"] = dict(
            SFs = SFs,
            distances_CC = distances_CC,
            separations = separations,
            collisions = collisions,
            # offset_md = mds,
            # offset_n = n,
            # offset_e = e,
            # offset_tvd = tvd
        )

        if well == "10 - well":
            kop_index = 31
        else:
            kop_index = 0

        data = get_standard_separation(data, reference, well, kop_index)

    # export the data to Excel
    with pd.ExcelWriter(f'data/output/output.xlsx') as writer:
        for well in data["wells"].keys():
            if well == "Reference well": continue
            df = pd.DataFrame(data["wells"][well]["results"])
            if well == "10 - well":
                df.insert(0, "MD", pd.DataFrame(np.array(data["wells"]["Reference well"]["MD"])[np.where(np.array(data["wells"]["Reference well"]["MD"]) >= 900)]))
            else:
                df.insert(0, "MD", pd.DataFrame(data["wells"]["Reference well"]["MD"]))
            df.to_excel(writer, sheet_name=f'{well} - output')


    print("Hello world")

        