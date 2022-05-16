try:
    from openpyxl import load_workbook
    OPENPYXL = True
except:
    OPENPYXL = False
from math import radians
import yaml
import os

"""
Code to extract error model data from standard ISCWSA Excel files and create
yaml files for use with the welleng.error module.

It's coded up this way so that when the error models are updated then simply
add the new workbook to the reference folder and add a new line to remake
the error model input data.

If new error functions are added then these will need to be coded into the
error module.
"""

PATH = os.path.dirname(__file__)

FILENAME = 'welleng/errors/error_codes.yaml'
CHARACTERS = [":", "[", "]"]

try:
    with open(FILENAME, 'r') as file:
        ec = yaml.full_load(file)
except Exception:
    ec = {}


def extract_data(filename):
    assert OPENPYXL, "ImportError: try pip install welleng[easy]"
    workbook = load_workbook(
        filename=filename,
        data_only=True
    )
    m = workbook['Model']

    e = extract_codes(m)
    h = extract_header(m)

    ed = dict(
        header=h,
        codes=e
    )

    return ed


def extract_header(workbook):
    m = workbook
    d = []

    for row in m.iter_rows(
        min_row=0,
        max_row=(),
        min_col=0,
        max_col=2,
        values_only=True
    ):
        if row[0] is None:
            continue
        else:
            d.append([
                row[0], row[1]
            ])

    h = make_header_dict(d)

    return h


def extract_codes(workbook):
    m = workbook
    d = []

    for row in m.iter_rows(
        min_row=4,
        max_row=(),
        min_col=4,
        max_col=(),
        values_only=True
    ):
        code = row[1]
        if code == "XCLL":
            code = "XCLA"
        function = row[3]
        magnitude = row[6]
        unit = row[7]
        propagation = row[8]
        d.append([
            code, function, magnitude, unit, propagation
        ])

    e = make_error_dict(d)

    return e


def make_error_dict(data):
    e = {}

    for c, f, m, u, p in data:
        e[c] = dict(
            function=f.replace('-', '_'),
            magnitude=m if 'deg' not in u else radians(m),
            unit=u.replace("deg", "rad"),
            propagation='random' if p == 'R' else 'systematic'
        )

    return e


def make_header_dict(data):
    h = {}

    for k, v in data:
        kk = remove_characters(k)
        h[kk] = v

    return h


def remove_characters(data, chars=CHARACTERS):
    for c in chars:
        data = data.replace(c, "")

    return data


def make_index(tool_code, data):
    fields = [
        'Application', 'Short Name', 'Long Name', 'OWSG Prefix', 'Revision No'
    ]
    temp = {
        f: data['header'][f] for f in fields
    }
    temp = {
        (k if k != 'Revision No' else 'Rev.'): v
        for k, v in temp.items()
    }
    filename = os.path.join(
        '',
        *[PATH, 'tool_index.yaml']
    )

    # read tool_index file
    with open(filename, 'r') as f:
        tool_index = yaml.safe_load(f)

    # add (or overwrite) new entry
    tool_index[tool_code] = temp

    # save new tool_index
    with open(filename, 'w') as f:
        yaml.dump(tool_index, f)


if __name__ == '__main__':
    # # extract rev 4 model
    tool_code = 'ISCWSA MWD Rev4'
    ec = extract_data(
        "reference/error-model-example-mwdrev4-iscwsa-1.xlsx"
    )

    # write data to yaml
    filename = os.path.join(
        '',
        *[PATH, 'tool_codes', f'{tool_code}.yaml']
    )

    with open(filename, 'w') as f:
        documents = yaml.dump(ec, f)

    # update index file
    make_index(tool_code, ec)


    # extract rev 5 model
    tool_code = 'ISCWSA MWD Rev5'
    ec = extract_data(
        "reference/error-model-example-mwdrev5-iscwsa-1.xlsx"
    )

    filename = os.path.join(
        '',
        *[PATH, 'tool_codes', f'{tool_code}.yaml']
    )

    with open(filename, 'w') as f:
        documents = yaml.dump(ec, f)

    make_index(tool_code, ec)

    print("Done")
