from openpyxl import load_workbook
from math import radians
import pandas as pd
import yaml
import os

"""
Code to extract error model data from standard ISCWSA Excel files and create
yaml files for use with the welleng.error module.

It's coded up this way so that when the error models are updated then simply
add the new workbook to the reference folder and add a new line to remake
the error model input data.

If new error functions are added then these will need to be coded into the
error module.
"""

PATH = os.path.dirname(__file__)
# PATH = os.path.join('', *[PATH, 'tool_codes'])

# FILENAME = 'welleng/errors/error_codes.yaml'
CHARACTERS = [":", "[", "]"]

# try:
#     with open(FILENAME, 'r') as file:
#         ec = yaml.full_load(file)
# except Exception:
#     ec = {}


def extract_data(file, sheet_name='Model'):
    # check if file is workbook otherwise treat as filename
    try:
        m = file[sheet_name]
    except TypeError:
        workbook = load_workbook(
            filename=file,
            data_only=True
        )
        m = workbook[sheet_name]

    e = extract_codes(m)
    h = extract_header(m)

    ed = dict(
        header=h,
        codes=e
    )

    return ed


def extract_header(worksheet):
    m = worksheet
    d = []
    tag = False
    temp = {}

    for row in m.iter_rows(
        min_row=0,
        max_row=(),
        min_col=0,
        max_col=2,
        values_only=True
    ):
        if tag:
            if row[0][0] == " ":
                if 'deg' in row[1]:
                    value = radians(float(row[1].split(" ")[0]))
                else:
                    value = row[1]
                temp[tag][row[0].lstrip().replace(":", "")] = value
                continue
            else:
                tag = False
        if row[0] is None:
            continue
        if "gyro" in row[0].lower():
            tag = row[0].replace(":", "")
            temp[tag] = {}
        else:
            d.append([
                row[0], row[1]
            ])

    h = make_header_dict(d)

    if bool(temp):
        h.update(temp)

    return h


def extract_codes(worksheet):
    m = worksheet
    d = []

    for row in m.iter_rows(
        min_row=4,
        max_row=(),
        min_col=4,
        max_col=(),
        values_only=True
    ):
        code = row[1]
        if code is None:
            continue
        if code == "XCLL":
            code = "XCLA"
        function = row[3]
        magnitude = row[6]
        unit = row[7]
        propagation = row[8]
        d.append([
            code, function, magnitude, unit, propagation
        ])

    e = make_error_dict(d)

    return e


def make_error_dict(data):
    e = {}

    for c, f, m, u, p in data:
        e[c] = dict(
            function=f.replace('-', '_'),
            magnitude=m if 'deg' not in u else radians(m),
            unit=u.replace("deg", "rad"),
            propagation='random' if p == 'R' else 'systematic'
        )

    return e


def make_header_dict(data):
    h = {}

    for k, v in data:
        kk = remove_characters(k)
        h[kk] = v

    return h


def remove_characters(data, chars=CHARACTERS):
    for c in chars:
        data = data.replace(c, "")

    return data


def open_workbook(filename):
    workbook = load_workbook(
        filename=filename,
        data_only=True
    )

    return workbook


def make_index(wb):
    d = {}

    for i, row in enumerate(wb['Index'].iter_rows(
        values_only=True
    )):
        if i == 0:
            fields = row
            continue
        if row[2].split('_')[-1] == 'Fl':
            owsg_prefix = f"{row[1]}_Fl"
        else:
            owsg_prefix = row[1]
        d[owsg_prefix] = {}
        for f, r in zip(fields, row):
            d[owsg_prefix][f] = r

    return d


def get_index(short_name, index_file):
    for k, v in index_file.items():
        if v['Short Name'] == short_name:
            return k

    return None


def get_short_names(index_file):
    try:
        with open(index_file, 'r') as f:
            index_file = yaml.safe_load(f)
    except TypeError:
        pass

    short_names = {
        k: v['Short Name']
        for k, v in index_file.items()
    }

    return short_names


if __name__ == '__main__':
    sheet_ignore_list = ['Sheet']
    filename = (
        f'reference/'
        f'toolgroup-owsg-a-rev-5-1-08-oct-2020-produced-22-oct-2020.xlsx'
    )

    wb = open_workbook(
        filename
    )

    temp = make_index(wb)
    filename = os.path.join(
        '',
        *[PATH, 'tool_index.yaml']
    )

    # open tool index file
    with open(filename, 'r') as f:
        tool_index = yaml.safe_load(f)

    # add the new tools
    for k, v in temp.items():
        tool_index[k] = v

    # write the updated file
    with open(filename, 'w') as f:
        yaml.dump(tool_index, f)

    short_names = get_short_names(tool_index)

    tool = get_index([*short_names.values()][10], tool_index)

    ec = {}

    # sheets are named by 'Short Name'
    for k, v in temp.items():
        sheet_name = v['Short Name']
        for s in sheet_ignore_list:
            if s in sheet_name:
                continue

        ec[k] = extract_data(
            file=wb, sheet_name=sheet_name
        )

    for k, v in ec.items():
        owsg_prefix = v['header']['OWSG Prefix']
        filename = os.path.join(
            '',
            *[PATH, 'tool_codes', f"{k}.yaml"]
        )
        with open(filename, 'w') as f:
            yaml.dump(ec[k], f)

    # wb = open_workbook(
    #     'reference/toolgroup-owsg-a-rev-5-1-08-oct-2020-produced-22-oct-2020.xlsx'
    # )



    # ec['iscwsa_mwd_rev4'] = extract_data(
    #     "toolgroup-owsg-a-rev-5-1-08-oct-2020-produced-22-oct-2020.xlsx"
    # )

    # ec['iscwsa_mwd_rev5'] = extract_data(
    #     "reference/error-model-example-mwdrev5-iscwsa-1.xlsx"
    # )

    # with open(
    #     os.path.join("", *[PATH, "tool_codes", "ISCWSA MWD Rev5.yaml"]),
    #     'w'
    # ) as f:
    #     documents = yaml.dump(ec, f)

    print("Done")
