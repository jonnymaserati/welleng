"""
Clean ingester for ISCWSA MWD error-model example workbooks.

Reads the Model, Wellpath, Validation, TOTALS and per-error-source sheets into
typed dataclasses so the data can be used as weight inputs, validation targets
or round-trip test fixtures.

Designed for the rev 5.1 example workbooks under ``data/iscwsa/rev5_11/`` but
the sheet layout is inherited from earlier revisions, so Rev4 workbooks parse
too (the extra P1/P2/P3 and formula columns are simply absent).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np


# Readme revision lines look like "Revision 5.11" (no colon). The
# ``Revision: 1.1`` line refers to the Readme's own format and is ignored.
_REVISION_RE = re.compile(r"^\s*Revision\s+(\d+)\.(\d+)\s*$")


def _latest_readme_revision(ws: Worksheet) -> str | None:
    """Scan column A of the Readme sheet and return the latest "Revision X.YY".

    Multiple revision lines form a version-history list; the highest numeric
    (major, minor) tuple wins. Returns ``None`` if no such line is found.
    """
    best: tuple[int, int] | None = None
    best_str: str | None = None
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        v = row[0]
        if v is None:
            continue
        m = _REVISION_RE.match(str(v))
        if not m:
            continue
        key = (int(m.group(1)), int(m.group(2)))
        if best is None or key > best:
            best = key
            best_str = str(v).strip()
    return best_str

try:
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    Worksheet = Any  # type: ignore


PROPAGATION_MAP = {
    "R": "random",
    "S": "systematic",
    "G": "global",
    "W": "within_pad",
}


@dataclass
class ErrorTerm:
    no: int
    code: str
    description: str
    wt_fn: str
    wt_fn_source: str
    type: str
    magnitude: float
    unit: str
    propagation: str  # raw single-letter code (R/S/G/W)
    p1: float | None = None
    p2: float | None = None
    p3: float | None = None
    wt_fn_comment: str | None = None
    depth_formula: str | None = None
    inc_formula: str | None = None
    az_formula: str | None = None
    sing_n_formula: str | None = None
    sing_e_formula: str | None = None
    sing_v_formula: str | None = None
    convert_deg_to_rad: bool = False

    @property
    def propagation_name(self) -> str:
        return PROPAGATION_MAP.get(self.propagation, self.propagation)

    @property
    def magnitude_si(self) -> float:
        """Magnitude with any 'deg' unit converted to radians."""
        if self.unit and "deg" in self.unit:
            return float(np.radians(self.magnitude))
        return float(self.magnitude)

    @property
    def unit_si(self) -> str:
        return self.unit.replace("deg", "rad") if self.unit else self.unit


@dataclass
class ModelHeader:
    owsg_prefix: str | None = None
    short_name: str | None = None
    long_name: str | None = None
    revision_no: Any = None
    revision_date: datetime | None = None
    revision_comment: str | None = None
    source: str | None = None
    application: str | None = None
    tool_type: str | None = None
    status: str | None = None
    checked: str | None = None
    approved: str | None = None
    notes: str | None = None
    extras: dict[str, Any] = field(default_factory=dict)


@dataclass
class Wellpath:
    well_name: str
    latitude_deg: float
    g: float
    btotal: float
    dip_deg: float
    declination_deg: float
    stations: np.ndarray  # (n, 5): md, inc_deg, azi_deg, tvd, toolface_deg
    depth_units: str = "m"         # "m" or "ft" as declared on the Wellpath sheet
    ft_to_m: float = 0.3048        # workbook's declared ft→m factor


@dataclass
class CovarianceStations:
    """Per-station covariance block. ``nev`` / ``hla`` columns ordered:

    NEV: [NN, EE, VV, NE, NV, EV]
    HLA: [HH, LL, AA, HL, HA, LA]
    """

    md: np.ndarray
    nev: np.ndarray
    hla: np.ndarray


@dataclass
class ErrorSourceTab:
    """Per-error-source sheet parsed into named numeric blocks keyed by md."""

    code: str
    md: np.ndarray
    blocks: dict[str, dict[str, np.ndarray]] = field(default_factory=dict)


@dataclass
class ValidationRow:
    md: float
    source: str
    diagnostic: np.ndarray  # len 6: NN EE VV NE NV EV
    calculated: np.ndarray  # len 6


@dataclass
class ISCWSAExample:
    path: Path
    revision: str | None  # workbook-level revision from Readme!A9
    header: ModelHeader
    terms: list[ErrorTerm]
    wellpath: Wellpath
    totals: CovarianceStations
    validation: list[ValidationRow]
    sources: dict[str, ErrorSourceTab]


# ---------------------------------------------------------------------------
# low-level helpers
# ---------------------------------------------------------------------------


def _require_openpyxl() -> None:
    if not _HAS_OPENPYXL:
        raise ImportError(
            "openpyxl is required for ISCWSA ingestion; "
            "install with `pip install welleng[easy]` or `pip install openpyxl`"
        )


def _rows(ws: Worksheet, **kw) -> list[tuple]:
    return list(ws.iter_rows(values_only=True, **kw))


def _to_float(v: Any) -> float:
    if v is None:
        return float("nan")
    try:
        return float(v)
    except (TypeError, ValueError):
        return float("nan")


def _merged_header_spans(ws: Worksheet, header_row: int) -> dict[int, tuple[str, int]]:
    """Return ``{start_col_idx: (label, span)}`` for merged cells in ``header_row``.

    Column indices are 0-based; ``header_row`` is 1-based (openpyxl convention).
    Unmerged labelled cells are included as span=1.
    """
    spans: dict[int, tuple[str, int]] = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row == header_row == mr.max_row:
            label = ws.cell(row=mr.min_row, column=mr.min_col).value
            if label is not None:
                spans[mr.min_col - 1] = (str(label), mr.max_col - mr.min_col + 1)

    # Pick up single-cell labels (not in a merged range)
    for cell in ws[header_row]:
        idx = cell.column - 1
        if cell.value is not None and idx not in spans:
            # ensure this cell isn't inside a merged range already accounted for
            inside_existing = any(
                start <= idx < start + span for start, (_, span) in spans.items()
            )
            if not inside_existing:
                spans[idx] = (str(cell.value), 1)
    return spans


# ---------------------------------------------------------------------------
# sheet parsers
# ---------------------------------------------------------------------------


_HEADER_KEYMAP = {
    "OWSG Prefix": "owsg_prefix",
    "Short Name": "short_name",
    "Long Name": "long_name",
    "Revision No": "revision_no",
    "Revision Date": "revision_date",
    "Revision Comment": "revision_comment",
    "Source": "source",
    "Application": "application",
    "Tool Type": "tool_type",
    "Status": "status",
    "Checked": "checked",
    "Approved": "approved",
    "Notes": "notes",
}

_TRIM = str.maketrans("", "", ":[]")


def parse_model_sheet(ws: Worksheet) -> tuple[ModelHeader, list[ErrorTerm]]:
    header = ModelHeader()
    terms: list[ErrorTerm] = []

    for row in _rows(ws):
        key = row[0]
        if key:
            norm = str(key).translate(_TRIM).strip()
            attr = _HEADER_KEYMAP.get(norm)
            if attr is not None:
                setattr(header, attr, row[1])
            else:
                header.extras[norm] = row[1]

        # Error-term table starts at col index 3 (col D = 'No')
        no = row[3] if len(row) > 3 else None
        code = row[4] if len(row) > 4 else None
        if isinstance(no, (int, float)) and code:
            terms.append(
                ErrorTerm(
                    no=int(no),
                    code=str(code).strip(),
                    description=str(row[5]) if row[5] is not None else "",
                    wt_fn=str(row[6]) if row[6] is not None else "",
                    wt_fn_source=str(row[7]) if row[7] is not None else "",
                    type=str(row[8]) if row[8] is not None else "",
                    magnitude=_to_float(row[9]),
                    unit=str(row[10]) if row[10] is not None else "",
                    propagation=str(row[11]).strip() if row[11] is not None else "",
                    p1=row[12] if len(row) > 12 else None,
                    p2=row[13] if len(row) > 13 else None,
                    p3=row[14] if len(row) > 14 else None,
                    wt_fn_comment=row[15] if len(row) > 15 else None,
                    depth_formula=row[16] if len(row) > 16 else None,
                    inc_formula=row[17] if len(row) > 17 else None,
                    az_formula=row[18] if len(row) > 18 else None,
                    sing_n_formula=row[19] if len(row) > 19 else None,
                    sing_e_formula=row[20] if len(row) > 20 else None,
                    sing_v_formula=row[21] if len(row) > 21 else None,
                    convert_deg_to_rad=bool(row[22]) if len(row) > 22 and row[22] is not None else False,
                )
            )
    return header, terms


def parse_wellpath_sheet(ws: Worksheet) -> Wellpath:
    # Reference params live in col A (label) / col B (value) of the top rows.
    refs: dict[str, float] = {}
    well_name = ""
    depth_units = "m"
    ft_to_m = 0.3048
    for row in _rows(ws, max_row=12):
        if not row or row[0] is None:
            continue
        label = str(row[0]).strip()
        if label == "Well":
            well_name = str(row[1]) if row[1] is not None else ""
            continue
        if label.lower().startswith("depth units"):
            if row[1] is not None:
                depth_units = str(row[1]).strip().lower()
            continue
        if label.lower().startswith("feet to metres"):
            if row[1] is not None and isinstance(row[1], (int, float)):
                ft_to_m = float(row[1])
            continue
        # Reference params are formatted as "Name (unit)"
        key = label.split("(")[0].strip().lower()
        if row[1] is not None and isinstance(row[1], (int, float)):
            refs[key] = float(row[1])

    # Survey table starts at col E (idx 4), with header on row 2 (idx 1).
    stations: list[list[float]] = []
    for row in _rows(ws, min_row=3):  # 1-based -> skip header rows
        md = row[4] if len(row) > 4 else None
        if md is None or not isinstance(md, (int, float)):
            continue
        inc = row[5] if len(row) > 5 else 0
        azi = row[6] if len(row) > 6 else 0
        tvd = row[7] if len(row) > 7 else 0
        tf = row[8] if len(row) > 8 else None
        stations.append([
            float(md),
            _to_float(inc),
            _to_float(azi),
            _to_float(tvd),
            _to_float(tf),
        ])

    return Wellpath(
        well_name=well_name,
        latitude_deg=refs.get("latitude", float("nan")),
        g=refs.get("g", float("nan")),
        btotal=refs.get("btotal", float("nan")),
        dip_deg=refs.get("dip", float("nan")),
        declination_deg=refs.get("declination", float("nan")),
        stations=np.asarray(stations, dtype=float),
        depth_units=depth_units,
        ft_to_m=ft_to_m,
    )


# NEV / HLA column ordering used throughout the workbook
_NEV_COLS = ("NN", "EE", "VV", "NE", "NV", "EV")
_HLA_COLS = ("HH", "LL", "AA", "HL", "HA", "LA")


def _extract_cov_columns(
    header_row: tuple, data: list[list[Any]], labels: tuple[str, ...]
) -> np.ndarray:
    idxs = [header_row.index(lbl) if lbl in header_row else None for lbl in labels]
    rows = []
    for r in data:
        rows.append([_to_float(r[i]) if i is not None and i < len(r) else float("nan") for i in idxs])
    return np.asarray(rows, dtype=float)


def parse_totals_sheet(ws: Worksheet) -> CovarianceStations:
    rows = _rows(ws)
    # Row 1 (idx 1) is the column-name header (Md, NN, EE, ...)
    colnames = rows[1]
    data = [list(r) for r in rows[2:] if r and isinstance(r[0], (int, float))]
    md_idx = colnames.index("Md") if "Md" in colnames else 0
    md = np.asarray([r[md_idx] for r in data], dtype=float)
    # The NEV block appears once and the HLA block once; find the first
    # occurrence of each label.
    nev = _extract_cov_columns(colnames, data, _NEV_COLS)
    hla = _extract_cov_columns(colnames, data, _HLA_COLS)
    return CovarianceStations(md=md, nev=nev, hla=hla)


def parse_validation_sheet(ws: Worksheet) -> list[ValidationRow]:
    rows = _rows(ws)
    if len(rows) < 3:
        return []
    colnames = list(rows[1])
    # There are two blocks of NN..EV: diagnostic (first 6) and calculated
    # (second 6). Locate by index.
    nev_positions = [i for i, c in enumerate(colnames) if c in _NEV_COLS]
    # Take first 6 for diagnostic, next 6 for calculated
    diag_idx = nev_positions[:6]
    calc_idx = nev_positions[6:12]
    out: list[ValidationRow] = []
    md_idx = colnames.index("MD") if "MD" in colnames else 0
    src_idx = colnames.index("Source") if "Source" in colnames else 1
    for r in rows[2:]:
        md = r[md_idx]
        src = r[src_idx]
        if md is None or src is None:
            continue
        diag = np.asarray([_to_float(r[i]) for i in diag_idx], dtype=float)
        calc = np.asarray([_to_float(r[i]) for i in calc_idx], dtype=float)
        out.append(ValidationRow(float(md), str(src), diag, calc))
    return out


def parse_error_source_sheet(ws: Worksheet, code: str) -> ErrorSourceTab:
    """Generic parser for per-error-source sheets (XCLA, DRFR, ABZ, ...).

    Row 1 holds merged block labels ('dpde', 'e', 'estar', 'Sigma e ...',
    'Covariance', etc). Row 2 holds per-column sub-labels. Data follows.
    """
    spans = _merged_header_spans(ws, header_row=1)
    rows = _rows(ws)
    if len(rows) < 3:
        return ErrorSourceTab(code=code, md=np.empty(0))
    subcols = list(rows[1])
    data = [list(r) for r in rows[2:] if r and isinstance(r[0], (int, float))]

    md = np.asarray([r[0] for r in data], dtype=float)

    blocks: dict[str, dict[str, np.ndarray]] = {}
    # Disambiguate repeated block labels (e.g. two 'Covariance' groups)
    seen: dict[str, int] = {}
    for start in sorted(spans):
        label, span = spans[start]
        if label == subcols[0]:  # skip the 'Md' column label if it shows up here
            continue
        count = seen.get(label, 0)
        name = label if count == 0 else f"{label}_{count + 1}"
        seen[label] = count + 1

        block: dict[str, np.ndarray] = {}
        for offset in range(span):
            col_idx = start + offset
            sub = subcols[col_idx] if col_idx < len(subcols) else None
            if sub is None:
                continue
            arr = np.asarray(
                [_to_float(r[col_idx]) if col_idx < len(r) else float("nan") for r in data],
                dtype=float,
            )
            block[str(sub)] = arr
        if block:
            blocks[name] = block
    return ErrorSourceTab(code=code, md=md, blocks=blocks)


# ---------------------------------------------------------------------------
# top-level
# ---------------------------------------------------------------------------

_NON_SOURCE_SHEETS = frozenset({
    "Readme", "Model", "Wellpath", "Validation", "CovChart", "TOTALS", "drdp",
})


def load_iscwsa_example(path: str | Path) -> ISCWSAExample:
    """Parse an ISCWSA MWD example workbook into typed dataclasses."""
    _require_openpyxl()
    path = Path(path)
    wb = load_workbook(filename=str(path), data_only=True)

    revision = None
    if "Readme" in wb.sheetnames:
        revision = _latest_readme_revision(wb["Readme"])

    header, terms = parse_model_sheet(wb["Model"])
    wellpath = parse_wellpath_sheet(wb["Wellpath"])
    totals = parse_totals_sheet(wb["TOTALS"])
    validation = parse_validation_sheet(wb["Validation"]) if "Validation" in wb.sheetnames else []

    term_codes = {t.code for t in terms}
    n_stations = wellpath.stations.shape[0]
    sources: dict[str, ErrorSourceTab] = {}
    for name in wb.sheetnames:
        if name in _NON_SOURCE_SHEETS:
            continue
        # Only treat a sheet as a per-source tab if it matches an error term
        # code. Anything else is an auxiliary sheet we don't understand yet.
        if name in term_codes:
            tab = parse_error_source_sheet(wb[name], name)
            # Per-source tabs sometimes carry trailing template rows beyond the
            # wellpath length; truncate to the wellpath station count so all
            # arrays are station-aligned with TOTALS/Wellpath.
            if n_stations and tab.md.shape[0] > n_stations:
                tab.md = tab.md[:n_stations]
                tab.blocks = {
                    g: {c: arr[:n_stations] for c, arr in cols.items()}
                    for g, cols in tab.blocks.items()
                }
            sources[name] = tab

    return ISCWSAExample(
        path=path,
        revision=revision,
        header=header,
        terms=terms,
        wellpath=wellpath,
        totals=totals,
        validation=validation,
        sources=sources,
    )


def to_legacy_codes_dict(example: ISCWSAExample) -> dict[str, Any]:
    """Return the old ``extract_iscwsa_codes`` yaml shape for back-compat.

    Produces ``{'header': {...}, 'codes': {code: {function, magnitude, unit,
    propagation}}}`` with deg→rad conversion matched to the original behaviour.
    """
    header_out: dict[str, Any] = {}
    for k, v in {
        "OWSG Prefix": example.header.owsg_prefix,
        "Short Name": example.header.short_name,
        "Long Name": example.header.long_name,
        "Revision No": example.header.revision_no,
        "Revision Date": example.header.revision_date,
        "Revision Comment": example.header.revision_comment,
        "Source": example.header.source,
        "Application": example.header.application,
        "Tool Type": example.header.tool_type,
        "Status": example.header.status,
        "Checked": example.header.checked,
        "Approved": example.header.approved,
        "Notes": example.header.notes,
    }.items():
        if v is not None:
            header_out[k] = v

    codes: dict[str, dict[str, Any]] = {}
    for t in example.terms:
        code = "XCLA" if t.code == "XCLL" else t.code
        codes[code] = {
            "function": t.wt_fn.replace("-", "_"),
            "magnitude": t.magnitude_si,
            "unit": t.unit_si,
            "propagation": PROPAGATION_MAP.get(t.propagation, "systematic"),
        }
    return {"header": header_out, "codes": codes}
