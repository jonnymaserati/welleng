import numbers
import numpy as np
from .survey import Survey

try:
    from openpyxl import load_workbook
    OPENPYXL = True
except ImportError:
    OPENPYXL = False


def get_standard_data(filename):
    # import data from Excel
    assert OPENPYXL, "ImportError: try pip install welleng[easy]"
    workbook = load_workbook(filename, data_only=True)
    sheets = workbook.sheetnames

    # extract the sheets with the input data
    wells = []
    wells.extend([well for well in list(sheets) if well.split()[-1] == "well"])

    data = {
        "acr": {},
        "wells": {},
    }

    for i, well in enumerate(wells):
        sheet = workbook[well]
        data = get_well_data(well, sheet, data)
        if well == "Reference well":
            data = acr_setup(sheet, data)
        else:
            sheet = workbook[f"{well[:5]}clearance"]
            data = get_clearance_data(well, sheet, data)

    return data


def get_well_data(well, sheet, data):
    temp = dict(
        MD=[],
        IncDeg=[],
        AziDeg=[],
        TVD=[],
        N=[],
        E=[],
        sigmaH=[],
        sigmaL=[],
        sigmaA=[],
    )

    for row in sheet.iter_rows(
        min_row=17,
        max_row=(),
        min_col=2,
        max_col=10,
    ):
        if isinstance(row[0].value, numbers.Real):
            temp["MD"].append(row[0].value)
            temp["IncDeg"].append(row[1].value)
            temp["AziDeg"].append(row[2].value)
            temp["TVD"].append(row[3].value)
            temp["N"].append(row[4].value)
            temp["E"].append(row[5].value)
            temp["sigmaH"].append(row[6].value)
            temp["sigmaL"].append(row[7].value)
            temp["sigmaA"].append(row[8].value)

    data["wells"][f"{well}"] = temp

    return data


def acr_setup(sheet, data):
    data["acr"]["Sm"] = sheet["I4"].value
    data["acr"]["sigmapa"] = sheet["I5"].value
    data["acr"]["k"] = sheet["I6"].value
    data["acr"]["reference_h_and_c"] = sheet["I7"].value
    data["acr"]["offset_h_and_c"] = sheet["I8"].value

    return data


def make_survey(data, well):
    start_nev = data["wells"]["offset"]["mesh"].NEVs[0]
    y, x, z = start_nev
    start_xyz = np.array([x, y, z])
    return Survey(
        md=data["wells"][well]["MD"],
        inc=data["wells"][well]["IncDeg"],
        azi=data["wells"][well]["AziDeg"],
        start_nev=start_nev,
        start_xyz=start_xyz
    )


def get_clearance_data(well, sheet, data):
    sf = []
    for row in sheet.iter_rows(
        min_row=5,
        max_row=(),
        min_col=15,
        max_col=16,
    ):
        sf.append(row[0].value)

    data["wells"][f"{well}"]['SF'] = sf

    return data


def import_iscwsa_collision_data(filename):
    data = get_standard_data(filename)

    return data


if __name__ == "__main__":
    filename = (
        "reference/standard-set-of-wellpaths"
        "-for-evaluating-clearance-scenarios-r4-17-may-2017.xlsm"
    )
    data = import_iscwsa_collision_data(filename)
