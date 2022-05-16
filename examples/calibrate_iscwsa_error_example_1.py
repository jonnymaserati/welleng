'''
Import the standard ISCWSA well surveys and generate the error ellipses using
the ISCWSA MWD Rev4 model. Then print the covariance matrix of the final survey
station so that the results can be compared with the diagnostic data curated
by ISCWSA.

author: Jonny Corcutt
email: jonnycorcutt@gmail.com
date: 29-09-2021
'''

import welleng as we

from openpyxl import load_workbook

# import the ISCWSA standard Excel data
print("Importing data...")
try:
    workbook = load_workbook(
        filename="examples/data/error-model-example-mwdrev4-iscwsa-1.xlsx",
        data_only=True
    )
except: # noqa E722
    print(
        "Make sure you have a local copy of ISCWSA's Excel file and have"
        "updated the location in the code."
    )

sheets = workbook.sheetnames
model = workbook['Model']
wellpath = workbook['Wellpath']

sh = we.survey.SurveyHeader(
    latitude=wellpath['B2'].value,
    G=wellpath['B3'].value,
    b_total=wellpath['B4'].value,
    dip=wellpath['B5'].value,
    declination=wellpath['B6'].value,
    convergence=wellpath['B7'].value,
    azi_reference=(
        'true' if wellpath['B8'].value == 'TRUE'
        else 'grid' if wellpath['B8'].value == 'GRID'
        else 'magnetic'
    ),
    depth_unit=wellpath['B9'].value,
    vertical_inc_limit=wellpath['B10'].value,
)

md = []
inc = []
azi = []

for row in wellpath.iter_rows(
    min_row=3,
    max_row=(),
    min_col=5,
):
    md.append(row[0].value)
    inc.append(row[1].value)
    azi.append(row[2].value)

s = we.survey.Survey(
    md=md,
    inc=inc,
    azi=azi,
    header=sh,
    error_model="ISCWSA MWD Rev4"
)

err0 = s.err

cov_nev0 = err0.errors.cov_NEVs.T

# print final covariance matrix for each tool
print("Tool errors at well TD:")
for tool, e in err0.errors.errors.items():
    print(
        f'{tool}:\n{e.cov_NEV.T[-1]}'
    )
print(
        f'TOTAL:\n{cov_nev0[-1]}'
    )

input("Press Enter to end...")
