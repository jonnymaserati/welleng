import json
from openpyxl import load_workbook

# Short script to extract and serialize the validation data from the
# ISCWSA error model spreadsheet. The intent is to save on test time versus
# having to extract the data from Excel workbook each time.
# This code will remain in the repo to demonstrate where the data is
# derived from and how it's been extracted.

# import workbook
filename = "examples/data/error-model-example-mwdrev4-iscwsa-1.xlsx"
workbook = load_workbook(filename, data_only=True)

validation_data = workbook["Validation"]

# initiate lists
md, source, nn, ee, vv, ne, nv, ev = [], [], [], [], [], [], [], []
data = [
    md, source, nn, ee, vv, ne, nv, ev
]

# extract data
for row in validation_data.iter_rows(
    min_row=3,
    max_row=(),
    min_col=0,
    max_col=8
):
    if row[0].value is None:
        continue
    for i, d in enumerate(data):
        d.append(row[i].value)

# convert to dictionary
vd = {}
headers = [
    'md', 'source', 'nn', 'ee', 'vv', 'ne', 'nv', 'ev'
]
for i, h in enumerate(headers):
    vd[h] = data[i]

# serialize data into file:
json.dump(vd, open("test/test_data/error_mwdrev4_iscwsa_validation.json", 'w'))
