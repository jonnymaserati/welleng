import json

from openpyxl import load_workbook

from welleng.survey import SurveyHeader


# Short script to extract and serialize the well data from the
# ISCWSA error model spreadsheet. The intent is to save on test time versus
# having to extract the data from Excel workbook each time.
# This code will remain in the repo to demonstrate where the data is
# derived and how it's been extracted.
def extract_well_data(input_file, output_file):
    workbook = open_workbook(input_file)
    well_data = get_survey_data(workbook)
    well_data = get_validation_data(workbook, well_data)
    save_well_data(well_data, output_file)


# import the ISCWSA standard Excel data
def open_workbook(filename):
    workbook = load_workbook(
        filename=filename,
        data_only=True
    )
    return workbook


def get_validation_data(workbook, well_data):
    validation_data = workbook["Validation"]

    # initiate lists
    vd = {}
    md, source, nn, ee, vv, ne, nv, ev = [], [], [], [], [], [], [], []
    data = [
        md, source, nn, ee, vv, ne, nv, ev
    ]

    # extract data
    for row in validation_data.iter_rows(
        min_row=3,
        max_row=(),
        min_col=0,
        max_col=8
    ):
        if row[0].value is None:
            continue
        for i, d in enumerate(data):
            d.append(row[i].value)

    # convert to dictionary
    headers = [
        'md', 'source', 'nn', 'ee', 'vv', 'ne', 'nv', 'ev'
    ]
    for i, h in enumerate(headers):
        vd[h] = data[i]

    well_data['vd'] = vd

    return well_data


def get_survey_header(workbook):
    wellpath = workbook['Wellpath']

    sh = SurveyHeader(
        latitude=wellpath['B2'].value,
        G=wellpath['B3'].value,
        b_total=wellpath['B4'].value,
        dip=wellpath['B5'].value,
        declination=wellpath['B6'].value,
        convergence=wellpath['B7'].value,
        azi_reference=(
            'true' if wellpath['B8'].value == 'TRUE'
            else 'grid' if wellpath['B8'].value == 'GRID'
            else 'magnetic'
        ),
        depth_unit=wellpath['B9'].value,
        vertical_inc_limit=wellpath['B10'].value
    )
    return sh


def get_survey_data(workbook):
    wellpath = workbook['Wellpath']

    sh = get_survey_header(workbook)

    md = []
    inc = []
    azi = []

    for row in wellpath.iter_rows(
        min_row=3,
        max_row=(),
        min_col=5,
        max_col=9
    ):
        if row[0].value is None:
            continue
        md.append(row[0].value)
        inc.append(row[1].value)
        azi.append(row[2].value)

    # convert required data to dictionary
    well_data = {}
    well_data['header'] = vars(sh)
    well_data['survey'] = dict(
        md=md,
        inc=inc,
        azi=azi
    )

    return well_data


def save_well_data(well_data, filename):
    # serialize data into file:
    json.dump(
        well_data,
        open(filename, 'w')
    )


if __name__ == "__main__":
    # generate iscwsa_mwd_rev4_1 data
    extract_well_data(
        input_file="reference/error-model-example-mwdrev4-iscwsa-1.xlsx",
        output_file="test/test_data/error_mwdrev4_1_iscwsa_data.json"
    )

    # generate iscwsa_mwd_rev5_1 data
    extract_well_data(
        input_file="reference/error-model-example-mwdrev5-iscwsa-1.xlsx",
        output_file="test/test_data/error_mwdrev5_1_iscwsa_data.json"
    )
