from openpyxl import load_workbook
import json
from welleng.survey import SurveyHeader

# Short script to extract and serialize the well data from the
# ISCWSA error model spreadsheet. The intent is to save on test time versus
# having to extract the data from Excel workbook each time.
# This code will remain in the repo to demonstrate where the data is
# derived and how it's been extracted.

# import the ISCWSA standard Excel data
workbook = load_workbook(
    filename="examples/data/error-model-example-mwdrev4-iscwsa-1.xlsx",
    data_only=True
)
# model = workbook['Model']
wellpath = workbook['Wellpath']

sh = SurveyHeader(
    latitude=wellpath['B2'].value,
    G=wellpath['B3'].value,
    b_total=wellpath['B4'].value,
    dip=wellpath['B5'].value,
    declination=wellpath['B6'].value,
    convergence=wellpath['B7'].value,
    azi_reference=(
        'true' if wellpath['B8'].value == 'TRUE'
        else 'grid' if wellpath['B8'].value == 'GRID'
        else 'magnetic'
    ),
    depth_unit=wellpath['B9'].value,
    vertical_inc_limit=wellpath['B10'].value
)

md = []
inc = []
azi = []

for row in wellpath.iter_rows(
    min_row=3,
    max_row=(),
    min_col=5,
    max_col=9
):
    md.append(row[0].value)
    inc.append(row[1].value)
    azi.append(row[2].value)

# convert required data to dictionary
well_data = {}
well_data['header'] = vars(sh)
well_data['survey'] = dict(
    md=md,
    inc=inc,
    azi=azi
)

# serialize data into file:
json.dump(
    well_data,
    open("test/test_data/error_mwdrev4_iscwsa_well_data.json", 'w')
)
